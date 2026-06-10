# core.student.views

import json
import random
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404, HttpResponseNotAllowed
from django.views import View
from django.core.paginator import Paginator
from django.db.models import Q
from datetime import timedelta
from decimal import Decimal
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from core.models import Student, Assignment, AssignmentSubmission, Teacher, CustomUser, Message
from core.models import Assessment, AssessmentSubmission, Exam, ExamSubmission
from .forms import StudentRegistrationForm, MessageForm, ReplyForm
from core.assignment.forms import AssignmentSubmissionForm
from core.utils import build_question_result, extract_submission_answer, question_answer_is_correct


# Student Views

def StudentRegisterView(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('student_dashboard')
    else:
        form = StudentRegistrationForm()
    return render(request, 'student/register.html', {'form': form})

def StudentProfileView(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('student_dashboard')
    else:
        form = StudentRegistrationForm(instance=request.user)
    return render(request, 'student/profile.html', {'form': form})

@login_required
def student_dashboard(request):
    if not request.user.role == 'student':
        return redirect('login')
    # Add any data for students
    return render(request, 'student_dashboard.html')
    
class StudentListView(View):
    template_name = 'student/student_list.html'

    def get(self, request, *args, **kwargs):
        from core.models import Class
        
        query = request.GET.get('q', '')
        status = request.GET.get('status', 'active')  # Default to 'active'
        per_page = int(request.GET.get('per_page', 20))  # Items per page selector

        # Filter students by status
        students = Student.objects.filter(status=status).select_related(
            'user', 'current_class', 'student_guardian__user'
        ).order_by('current_class', 'user__last_name')

        # Apply search filter
        if query:
            students = students.filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__email__icontains=query) |
                Q(LSA_number__icontains=query) |
                Q(current_class__name__icontains=query) |
                Q(student_guardian__user__first_name__icontains=query) |
                Q(student_guardian__user__last_name__icontains=query)
            )

        # Paginate students
        paginator = Paginator(students, per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Get stats for all statuses
        active_count = Student.objects.filter(status='active').count()
        dormant_count = Student.objects.filter(status='dormant').count()
        left_count = Student.objects.filter(status='left').count()
        total_count = active_count + dormant_count + left_count
        
        # Get all classes for bulk enrollment
        all_classes = Class.objects.all().order_by('order')

        return render(request, self.template_name, {
            'page_obj': page_obj,
            'active_tab': status,
            'active_count': active_count,
            'dormant_count': dormant_count,
            'left_count': left_count,
            'total_count': total_count,
            'all_classes': all_classes,
            'per_page': per_page,
        })

class BulkUpdateStudentsView(View):
    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        selected_students = request.POST.getlist("selected_students")

        if not selected_students:
            messages.error(request, "No students selected for the bulk action.")
            return redirect("student_list")

        # FIXED: Remove status filter to allow updates on all students
        students = Student.objects.filter(user__id__in=selected_students)

        if action == "promote":
            promoted_count = 0
            for student in students:
                if student.current_class and student.current_class.next_class:
                    student.current_class = student.current_class.next_class()
                    student.save()
                    promoted_count += 1
            messages.success(request, f"{promoted_count} student(s) have been promoted.")
        
        elif action == "demote":
            demoted_count = 0
            for student in students:
                if student.current_class and student.current_class.previous_class:
                    student.current_class = student.current_class.previous_class()
                    student.save()
                    demoted_count += 1
            messages.success(request, f"{demoted_count} student(s) have been demoted.")

        elif action == "mark_dormant":
            count = students.update(status="dormant")
            messages.success(request, f"{count} student(s) marked as dormant.")
        
        elif action == "mark_active":
            count = students.update(status="active")
            messages.success(request, f"{count} student(s) marked as active.")
        
        elif action == "mark_left":
            count = students.update(status="left")
            messages.success(request, f"{count} student(s) marked as left school.")
        
        elif action == "bulk_enroll":
            # Bulk enrollment to a class
            class_id = request.POST.get("target_class")
            if not class_id:
                messages.error(request, "Please select a class for enrollment.")
                return redirect("student_list")
            
            from core.models import Class
            target_class = get_object_or_404(Class, pk=class_id)
            
            # Check capacity if set
            if target_class.capacity:
                current_enrollment = target_class.enrolled_students.filter(status='active').count()
                available_slots = target_class.capacity - current_enrollment
                
                if len(selected_students) > available_slots:
                    messages.error(request, f"Cannot enroll {len(selected_students)} students. Only {available_slots} slots available in {target_class.name}.")
                    return redirect("student_list")
            
            enrolled_count = students.update(current_class=target_class)
            messages.success(request, f"{enrolled_count} student(s) enrolled to {target_class.name}.")
        
        else:
            messages.error(request, "Invalid action selected.")

        return redirect("student_list")


class StudentCreateView(View):
    template_name = 'student/student_form.html'

    def get(self, request, *args, **kwargs):
        form = StudentRegistrationForm()
        response = render(request, self.template_name, {'form': form})
        return response

    def post(self, request, *args, **kwargs):
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('student_list')
        else:
            print("Form is NOT valid. Errors: ", form.errors)
        return render(request, self.template_name, {'form': form})

class StudentUpdateView(View):
    template_name = 'student/student_form.html'

    def get(self, request, pk, *args, **kwargs):
        student = get_object_or_404(Student, pk=pk)
        form = StudentRegistrationForm(instance=student.user, student_instance=student, is_update=True)
        return render(request, self.template_name, {'form': form, 'is_update': True, 'student': student})

    def post(self, request, pk, *args, **kwargs):
        student = get_object_or_404(Student, pk=pk)
        form = StudentRegistrationForm(request.POST, request.FILES, instance=student.user, student_instance=student, is_update=True)

        if form.is_valid():
            form.save()  # This will save both user and student
            return redirect('student_detail', pk=student.pk)

        return render(request, self.template_name, {'form': form, 'is_update': True, 'student': student})

class StudentDetailView(View):
    template_name = 'student/student_detail.html'

    def get(self, request, pk, *args, **kwargs):
        student = get_object_or_404(Student, pk=pk)
        return render(request, self.template_name, {'student': student})

class StudentDeleteView(View):
    template_name = 'student/student_confirm_delete.html'

    def get(self, request, pk, *args, **kwargs):
        student = get_object_or_404(Student, pk=pk)
        return render(request, self.template_name, {'student': student})

    def post(self, request, pk, *args, **kwargs):
        student = get_object_or_404(Student, pk=pk)
        student.delete()
        return redirect('student_list')

# Example for export_students and student_reports
@login_required
def export_students(request):
    """Export student data to Excel or PDF based on user choice"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    
    export_format = request.GET.get('format', 'excel')  # Default to Excel
    status_filter = request.GET.get('status', '')  # Optional status filter
    
    # Get students based on filter
    students = Student.objects.select_related(
        'user', 'current_class', 'student_guardian__user'
    ).order_by('current_class', 'user__last_name')
    
    if status_filter:
        students = students.filter(status=status_filter)
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Header styling
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['LSA Number', 'First Name', 'Last Name', 'Email', 'Gender', 'Date of Birth', 
                   'Class', 'Guardian', 'Relationship', 'Status']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for student in students:
            ws.append([
                student.LSA_number,
                student.user.first_name,
                student.user.last_name,
                student.user.email,
                student.get_gender_display(),
                student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                student.current_class.name if student.current_class else 'Not Assigned',
                student.student_guardian.user.get_full_name() if student.student_guardian else 'N/A',
                student.relationship if student.relationship else '',
                student.get_status_display()
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response
        
    elif export_format == 'pdf':
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        # Table data
        data = [['LSA Number', 'Name', 'Email', 'Gender', 'DOB', 'Class', 'Guardian', 'Status']]
        
        for student in students:
            data.append([
                student.LSA_number,
                student.user.get_full_name(),
                student.user.email,
                student.get_gender_display(),
                student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                student.current_class.name if student.current_class else 'N/A',
                student.student_guardian.user.get_full_name() if student.student_guardian else 'N/A',
                student.get_status_display()
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="students_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    else:
        messages.error(request, "Invalid export format selected.")
        return redirect('student_list')

def student_reports(request):
    # Implement report generation logic here
    pass


# Helper to get client IP
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    return x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

# Helper to manage student/guardian logic to keep views DRY
def get_student_for_item(request, item, item_type_verbose, submit_url_name, template_path):
    """
    Handles student/guardian logic. Returns a student object or a redirect response.
    `item` can be an Assignment, Assessment, or Exam.
    """
    if hasattr(request.user, 'guardian'):
        guardian = request.user.guardian
        students = Student.objects.filter(student_guardian=guardian, current_class=item.class_assigned, status='active')
        if not students.exists():
            messages.error(request, "No students associated with this guardian are in this class.")
            return redirect('guardian_dashboard')
        
        student_id = request.GET.get('student_id')
        if student_id:
            return get_object_or_404(students, pk=student_id)
        elif students.count() == 1:
            return students.first()
        else:
            return render(request, template_path, {
                'students': students, 'item': item, 
                'item_type_verbose': item_type_verbose, 'submit_url_name': submit_url_name
            })
    else:
        return get_object_or_404(Student, user=request.user, current_class=item.class_assigned)


@login_required
def start_assignment(request, assignment_id):
    """STEP 1: Entry point. Creates submission, starts timer, gets IP, and redirects."""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    
    response = get_student_for_item(request, assignment)
    if not isinstance(response, Student):
        return response 
    student = response

    if AssignmentSubmission.objects.filter(assignment=assignment, student=student, is_completed=True).exists():
        return render(request, 'assignment/already_submitted.html', {"assignment": assignment, "student": student})

    submission, created = AssignmentSubmission.objects.get_or_create(
        assignment=assignment, student=student,
        defaults={'ip_address': get_client_ip(request)}
    )

    if not submission.started_at and assignment.duration:
        submission.started_at = timezone.now()
        submission.save()

    return redirect('take_assignment', submission_id=submission.id)


@login_required
def take_assignment(request, submission_id):
    """STEP 2: The main page where the student takes the assignment."""
    submission = get_object_or_404(AssignmentSubmission.objects.select_related('assignment', 'student__user'), id=submission_id)
    assignment = submission.assignment

    if submission.is_completed:
        messages.info(request, "This assignment has already been submitted.")
        return redirect('student_dashboard')

    end_time = None
    if assignment.duration and submission.started_at:
        end_time = submission.started_at + timedelta(minutes=assignment.duration)
        if timezone.now() > end_time:
            return render(request, 'assignment/assessment_due.html', {"assignment": assignment, "student": submission.student})

    questions_list = list(assignment.questions.all())
    if assignment.shuffle_questions:
        random.shuffle(questions_list)

    context = {
        'submission': submission, 'assignment': assignment, 'questions': questions_list,
        'student': submission.student, 'end_time_iso': end_time.isoformat() if end_time else None,
    }
    return render(request, 'assignment/take_assignment_form.html', context)


def _process_and_save_assignment_submission(request, submission_id):
    """Internal helper to process POST data for both regular and beacon submissions."""
    submission = get_object_or_404(AssignmentSubmission, id=submission_id)
    if submission.is_completed:
        return submission, False

    assignment = submission.assignment
    answers = {}
    auto_grade = 0
    
    for question in assignment.questions.all():
        answer_value = extract_submission_answer(request, question.id, question.question_type)
        answers[str(question.id)] = answer_value

        if question.question_type in ['SCQ', 'MCQ'] and question_answer_is_correct(question, answer_value):
            auto_grade += 1
    
    submission.answers = answers
    submission.grade = Decimal(auto_grade) # Awaiting manual grading for essays
    submission.is_completed = True
    submission.submitted_at = timezone.now()
    submission.save()
    return submission, True


@login_required
def submit_assignment(request, submission_id):
    """STEP 3: Handles the final, intentional submission from the student."""
    if request.method != "POST":
        return HttpResponseNotAllowed(['POST'])
    
    submission, was_processed = _process_and_save_assignment_submission(request, submission_id)
    if was_processed:
        messages.success(request, f"Assignment '{submission.assignment.title}' submitted successfully.")
    else:
        messages.warning(request, "Assignment was already submitted.")
    
    return redirect('guardian_dashboard' if hasattr(request.user, 'guardian') else 'student_dashboard')


@csrf_exempt
def autosubmit_beacon_assignment(request, submission_id):
    """STEP 4: Handles background submissions from browser closure."""
    if request.method != 'POST':
        return HttpResponse(status=204)
        
    submission, was_processed = _process_and_save_assignment_submission(request, submission_id)
    if was_processed:
        submission.force_submitted_at = timezone.now()
        submission.save()
    return HttpResponse(status=204)


@login_required
def start_assessment(request, assessment_id):

    assessment = get_object_or_404(Assessment, id=assessment_id, is_approved=True)

    if assessment.is_due:
        messages.error(request, f"The deadline for '{assessment.title}' has passed.")
        return render(request, 'assessment/assessment_due.html', {"assessment": assessment})

    # Get the student object
    response = get_student_for_item(request, assessment, 'Assessment', 'start_assessment', 'assessment/select_student.html')
    if not isinstance(response, Student): 
        return response
    student = response
 
    # Get all previous attempts for this student and assessment
    previous_attempts = AssessmentSubmission.objects.filter(
        assessment=assessment,
        student=student
    ).order_by('-attempt_number')
    
    latest_attempt = previous_attempts.first()
    attempt_count = previous_attempts.count()

    # Find the next attempt number
    next_attempt_number = (latest_attempt.attempt_number + 1) if latest_attempt else 1

    # CHECK 1: Have they already used all their attempts?
    if latest_attempt and latest_attempt.is_completed and attempt_count >= 3:
        messages.error(request, "You have used all available attempts for this assessment.")
        return render(request, 'assessment/already_submitted.html', {
            "assessment": assessment, "student": student, "latest_attempt": latest_attempt
        })

    # CHECK 2: Look for a pre-existing record for the next attempt.
    submission = AssessmentSubmission.objects.filter(
        assessment=assessment,
        student=student,
        attempt_number=next_attempt_number
    ).first()
    
    if not submission:
        # Before creating, one final check: if the last attempt was NOT completed, they can't start a new one.
        if latest_attempt and not latest_attempt.is_completed:
             messages.warning(request, "You have an incomplete attempt. Please finish it or contact your teacher.")
             # Redirect them back to their incomplete attempt
             return redirect('take_assessment', submission_id=latest_attempt.id)

        # All checks passed. Create the new attempt record.
        submission = AssessmentSubmission.objects.create(
            assessment=assessment,
            student=student,
            attempt_number=next_attempt_number,
            ip_address=get_client_ip(request)
        )

    # --- Timer and Redirect Logic (remains the same) ---
    if not submission.started_at and assessment.duration:
        submission.started_at = timezone.now()
        submission.save()

    return redirect('take_assessment', submission_id=submission.id)



@login_required
def take_assessment(request, submission_id):
    """STEP 2: The main page where the student takes the timed assessment."""
    submission = get_object_or_404(AssessmentSubmission.objects.select_related('assessment', 'student__user'), id=submission_id)
    assessment = submission.assessment

    if submission.is_completed:
        messages.info(request, "This assessment has already been submitted.")
        return redirect('student_dashboard')

    end_time = None
    if assessment.duration and submission.started_at:
        end_time = submission.started_at + timedelta(minutes=assessment.duration)
        if timezone.now() > end_time:
            return render(request, 'assessment/assessment_due.html', {"assessment": assessment, "student": submission.student})

    questions_list = list(assessment.questions.all())
    if assessment.shuffle_questions:
        random.shuffle(questions_list)
    
    processed_questions = []
    for q in questions_list:
        options = q.options_list()
        if assessment.shuffle_questions and q.question_type in ['SCQ', 'MCQ']:
            random.shuffle(options)
        processed_questions.append({'question': q, 'shuffled_options': options})

    context = {
        'submission': submission, 'assessment': assessment, 'questions_data': processed_questions,
        'student': submission.student, 'end_time_iso': end_time.isoformat() if end_time else None,
    }
    return render(request, 'assessment/take_assessment.html', context)


def _process_and_save_assessment_submission(request, submission_id):
    """Internal helper to process POST data for both regular and beacon submissions."""
    submission = get_object_or_404(AssessmentSubmission, id=submission_id)
    if submission.is_completed:
        return submission, False

    assessment = submission.assessment
    answers = {}
    score = Decimal('0.0')
    requires_manual_review = False

    for question in assessment.questions.all():
        answer_key = f'answer_{question.id}'
        submitted_answer = None
        if question.question_type == 'MCQ':
            submitted_answer = request.POST.getlist(answer_key)
        else:
            submitted_answer = request.POST.get(answer_key)
        
        answers[str(question.id)] = submitted_answer

        if question.question_type == 'ES':
            requires_manual_review = True
        elif question.is_option_correct(submitted_answer):
            score += Decimal(question.points) # Use points from OnlineQuestion
            
    submission.answers = answers
    submission.score = score
    submission.requires_manual_review = requires_manual_review
    submission.is_graded = not requires_manual_review
    submission.is_completed = True
    submission.submitted_at = timezone.now()
    submission.save() # The signal will fire after this save
    return submission, True


@login_required
def submit_assessment(request, submission_id):
    """STEP 3: Handles the final, intentional submission."""
    if request.method != "POST":
        return HttpResponseNotAllowed(['POST'])
    
    submission, was_processed = _process_and_save_assessment_submission(request, submission_id)
    if was_processed:
        messages.success(request, f"Assessment '{submission.assessment.title}' submitted successfully.")
    else:
        messages.warning(request, "Assessment was already submitted.")
    
    return redirect('guardian_dashboard' if hasattr(request.user, 'guardian') else 'student_dashboard')


@csrf_exempt
def autosubmit_beacon_assessment(request, submission_id):
    """STEP 4: Handles background submissions from browser closure."""
    if request.method != 'POST':
        return HttpResponse(status=204)
        
    submission, was_processed = _process_and_save_assessment_submission(request, submission_id)
    if was_processed:
        submission.force_submitted_at = timezone.now()
        submission.save()
    return HttpResponse(status=204)


@login_required
def start_exam(request, exam_id):

    exam = get_object_or_404(Exam, id=exam_id, is_approved=True)

    if exam.is_due:
        messages.error(request, f"The deadline for '{exam.title}' has passed.")
        return render(request, 'exam/exam_due.html', {"exam": exam})

    response = get_student_for_item(request, exam, 'Exam', 'start_exam', 'exam/select_student.html')
    if not isinstance(response, Student): 
        return response
    student = response

    # --- 3-ATTEMPT RETAKE LOGIC ---
    previous_attempts = ExamSubmission.objects.filter(exam=exam, student=student).order_by('-attempt_number')
    latest_attempt = previous_attempts.first()
    attempt_count = previous_attempts.count()
    next_attempt_number = (latest_attempt.attempt_number + 1) if latest_attempt else 1

    if latest_attempt and latest_attempt.is_completed and attempt_count >= 3:
        messages.error(request, "You have used all available attempts for this exam.")
        return render(request, 'exam/already_submitted.html', {"exam": exam, "student": student, "latest_attempt": latest_attempt})

    submission = ExamSubmission.objects.filter(exam=exam, student=student, attempt_number=next_attempt_number).first()
    
    if not submission:
        if latest_attempt and not latest_attempt.is_completed:
             messages.warning(request, "You have an incomplete attempt. Please finish it or contact your teacher.")
             return redirect('take_exam', submission_id=latest_attempt.id)

        submission = ExamSubmission.objects.create(
            exam=exam, student=student, attempt_number=next_attempt_number, ip_address=get_client_ip(request)
        )

    # --- Timer and Redirect ---
    if not submission.started_at and exam.duration:
        submission.started_at = timezone.now()
        submission.save()

    return redirect('take_exam', submission_id=submission.id)


@login_required
def take_exam(request, submission_id):
    """STEP 2: The main page where the student takes the timed exam."""
    submission = get_object_or_404(ExamSubmission.objects.select_related('exam', 'student__user'), id=submission_id)
    exam = submission.exam

    if submission.is_completed:
        messages.info(request, "This exam has already been submitted.")
        return redirect('student_dashboard')

    end_time = None
    if exam.duration and submission.started_at:
        end_time = submission.started_at + timedelta(minutes=exam.duration)
        if timezone.now() > end_time:
            return render(request, 'exam/exam_due.html', {"exam": exam, "student": submission.student})

    questions_list = list(exam.questions.all())
    if exam.shuffle_questions:
        random.shuffle(questions_list)
    
    processed_questions = []
    for q in questions_list:
        options = q.options_list()
        if exam.shuffle_questions and q.question_type in ['SCQ', 'MCQ']:
            random.shuffle(options)
        processed_questions.append({'question': q, 'shuffled_options': options})

    context = {
        'submission': submission, 'exam': exam, 'questions_data': processed_questions,
        'student': submission.student, 'end_time_iso': end_time.isoformat() if end_time else None,
    }
    return render(request, 'exam/take_exam.html', context)


def _process_and_save_exam_submission(request, submission_id):
    """Internal helper to process POST data for both regular and beacon submissions."""
    submission = get_object_or_404(ExamSubmission, id=submission_id)
    if submission.is_completed:
        return submission, False

    exam = submission.exam
    answers = {}
    score = Decimal('0.0')
    requires_manual_review = False

    for question in exam.questions.all():
        answer_key = f'answer_{question.id}'
        submitted_answer = None
        if question.question_type == 'MCQ':
            submitted_answer = request.POST.getlist(answer_key)
        else:
            submitted_answer = request.POST.get(answer_key)
        
        answers[str(question.id)] = submitted_answer

        if question.question_type == 'ES':
            requires_manual_review = True
        elif question.is_option_correct(submitted_answer):
            score += Decimal(question.points) # Use points from OnlineQuestion
            
    submission.answers = answers
    submission.score = score
    submission.requires_manual_review = requires_manual_review
    submission.is_graded = not requires_manual_review
    submission.is_completed = True
    submission.submitted_at = timezone.now()
    submission.save() # The signal will fire after this save
    return submission, True


@login_required
def submit_exam(request, submission_id):
    """STEP 3: Handles the final, intentional submission."""
    if request.method != "POST":
        return HttpResponseNotAllowed(['POST'])
    
    submission, was_processed = _process_and_save_exam_submission(request, submission_id)
    if was_processed:
        messages.success(request, f"Exam '{submission.exam.title}' submitted successfully.")
    else:
        messages.warning(request, "Exam was already submitted.")
    
    return redirect('guardian_dashboard' if hasattr(request.user, 'guardian') else 'student_dashboard')


@csrf_exempt
def autosubmit_beacon_exam(request, submission_id):
    """STEP 4: Handles background submissions from browser closure."""
    if request.method != 'POST':
        return HttpResponse(status=204)
        
    submission, was_processed = _process_and_save_exam_submission(request, submission_id)
    if was_processed:
        submission.force_submitted_at = timezone.now()
        submission.save()
    return HttpResponse(status=204)


@login_required
def message_inbox(request):
    """
    A central, generic inbox for the logged-in user.
    Shows received message threads.
    """
    user = request.user

    conversations = Message.objects.filter(
        (Q(recipient=user) | Q(sender=user)),
        parent_message__isnull=True
    ).select_related('sender', 'recipient', 'student_context__user').prefetch_related('replies').distinct().order_by('-updated_at')

    Message.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    unread_count = conversations.filter(recipient=user, is_read=False).count()
    
    context = {
        'conversations': conversations,
        'unread_count': unread_count, 
        'page_title': "My Inbox"
    }

    return render(request, 'messaging/inbox.html', context)


@login_required
def compose_message(request):
    """
    A generic view for composing a message.
    The list of possible recipients is determined by the sender's role.
    Recipient and other context can be pre-selected via GET parameters.
    """
    sender = request.user
    
    # --- Determine the queryset of possible recipients based on the sender's role ---
    recipient_qs = CustomUser.objects.none() 
    student_context_qs = Student.objects.none() 

    if sender.role == 'student':
        # Students can message their teachers and admins
        student = getattr(sender, 'student', None) 
        if student:
            teacher_ids = Teacher.objects.filter(
                Q(teacherassignment__class_assigned=student.current_class) |
                Q(subjectassignment__class_assigned=student.current_class),
                status='active'
            ).values_list('user_id', flat=True)
            recipient_qs = CustomUser.objects.filter(Q(id__in=teacher_ids) | Q(role='admin')).exclude(pk=sender.pk).distinct()
            student_context_qs = Student.objects.filter(pk=student.pk, status='active')

    elif sender.role == 'teacher':
        # Teachers can message guardians of their students, colleagues, and admins
        teacher = getattr(sender, 'teacher', None)
        if teacher:
            assigned_classes = teacher.assigned_classes()
            students_taught = Student.objects.filter(current_class__in=assigned_classes, status='active')
            guardian_ids = students_taught.filter(student_guardian__isnull=False).values_list('student_guardian__user_id', flat=True)
            
            colleague_ids = Teacher.objects.exclude(pk=teacher.pk).values_list('user_id', flat=True)
            
            recipient_qs = CustomUser.objects.filter(
                Q(guardian__pk__in=guardian_ids) |
                Q(id__in=colleague_ids) |
                Q(role='admin')
            ).distinct()
            student_context_qs = students_taught

    elif sender.role == 'guardian':
        # Guardians can message teachers of their wards and admins
        guardian = getattr(sender, 'guardian', None)
        if guardian:
            wards = guardian.students.all()
            teacher_ids = Teacher.objects.filter(
                Q(teacherassignment__class_assigned__in=wards.values_list('current_class', flat=True)) |
                Q(subjectassignment__class_assigned__in=wards.values_list('current_class', flat=True)),
                status='active'
            ).values_list('user_id', flat=True)
            recipient_qs = CustomUser.objects.filter(Q(id__in=teacher_ids) | Q(role='admin')).distinct()
            student_context_qs = wards

    elif sender.role == 'admin':
        # Admins can message anyone
        recipient_qs = CustomUser.objects.all().exclude(pk=sender.pk)
        student_context_qs = Student.objects.filter(status='active')
    
    # --- Initialize variables for context ---
    recipient_preselected = None
    student_context = None
    initial_form_data = {}

    # --- Check for pre-selected data from GET parameters ---
    recipient_id = request.GET.get('recipient_id')
    student_context_id = request.GET.get('student_context_id')

    if recipient_id:
        try:
            # Fetch the pre-selected recipient object to display their name
            recipient_preselected = recipient_qs.get(pk=recipient_id)
            initial_form_data['recipient'] = recipient_preselected
        except CustomUser.DoesNotExist:
            messages.error(request, "The specified recipient is invalid or you are not allowed to message them.")
            return redirect('message_inbox')
    
    if student_context_id:
        try:
            student_context = student_context_qs.get(pk=student_context_id)
            initial_form_data['student_context'] = student_context
        except Student.DoesNotExist:
            pass

    # --- Handle Form Submission ---
    if request.method == 'POST':
        form = MessageForm(request.POST, recipient_queryset=recipient_qs, student_queryset=student_context_qs)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = sender
            message.save()
            messages.success(request, "Message sent successfully!")
            return redirect('message_inbox')
    else: # GET request
        form = MessageForm(initial=initial_form_data, recipient_queryset=recipient_qs, student_queryset=student_context_qs)
    
    context = {
        'form': form,
        'recipient_preselected': recipient_preselected, 
        'student_context': student_context, 
        'page_title': "Compose Message"
    }
    return render(request, 'messaging/compose_message.html', context)


@login_required
def message_thread(request, thread_id):
    # Fetch the parent message of the thread
    parent_message = get_object_or_404(Message, pk=thread_id, parent_message__isnull=True)

    # Authorization check: user must be the sender or recipient
    if request.user not in [parent_message.sender, parent_message.recipient]:
        raise Http404

    # Mark all messages in this thread as read by the current user
    Message.objects.filter(
        Q(pk=thread_id) | Q(parent_message_id=thread_id),
        recipient=request.user, is_read=False
    ).update(is_read=True)
    
    thread_messages = parent_message.replies.all().order_by('sent_at')

    # Setup form for a new reply
    reply_recipient = parent_message.sender if request.user == parent_message.recipient else parent_message.recipient
    
    if request.method == 'POST':
        reply_form = ReplyForm(request.POST)
        if reply_form.is_valid():
            reply = reply_form.save(commit=False)
            reply.sender = request.user
            reply.recipient = reply_recipient
            reply.parent_message = parent_message
            if parent_message.student_context: # Carry over the student context
                reply.student_context = parent_message.student_context
            reply.save()
            messages.success(request, "Reply sent.")
            return redirect('message_thread', thread_id=thread_id)
    else:
        reply_form = MessageForm(initial={'title': f"Re: {parent_message.title}"})

    context = {
        'parent_message': parent_message,
        'thread_messages': thread_messages,
        'reply_form': reply_form,
        'reply_recipient': reply_recipient,
    }
    return render(request, 'messaging/message_thread.html', context)


@login_required
def view_assignment_result(request, submission_id):
    submission = get_object_or_404(AssignmentSubmission.objects.select_related(
        'assignment__subject', 'student__user'
    ), id=submission_id)

    # Authorization check
    is_student = request.user == submission.student.user
    is_guardian = hasattr(request.user, 'guardian') and submission.student in request.user.guardian.students.all()

    if not (is_student or is_guardian):
        messages.error(request, "You are not authorized to view these results.")
        return redirect('home')

    assignment = submission.assignment
    questions = assignment.questions.all()
    
    # --- THE DEFENSIVE FIX: Handle incorrectly saved data ---
    student_answers = submission.answers
    if isinstance(student_answers, str):
        try:
            student_answers = json.loads(student_answers)
        except json.JSONDecodeError:
            student_answers = {} # Default to empty dict if the string is invalid

    # Now, `student_answers` is guaranteed to be a dictionary
    
    # Prepare data for the template
    results_data = []
    for question in questions:
        student_answer = student_answers.get(str(question.id), "Not Answered")
        results_data.append(build_question_result(question, student_answer))

    context = {
        'submission': submission,
        'assignment': assignment,
        'results_data': results_data, # Pass the processed data
    }
    return render(request, 'assignment/view_assignment_result.html', context)


@login_required
def view_assessment_result(request, submission_id):
    """
    This view correctly uses the unique submission_id to get a single object.
    """
    # --- THE FIX: Use get_object_or_404 with the unique submission_id ---
    submission = get_object_or_404(AssessmentSubmission.objects.select_related(
        'assessment__subject', 'student__user'
    ), id=submission_id)

    # Authorization check
    is_student = request.user == submission.student.user
    is_guardian = hasattr(request.user, 'guardian') and submission.student in request.user.guardian.students.all()

    if not (is_student or is_guardian):
        messages.error(request, "You are not authorized to view these results.")
        return redirect('home')

    assessment = submission.assessment
    questions = assessment.questions.all()
    
    # Defensively handle the 'answers' field
    student_answers = submission.answers
    if isinstance(student_answers, str):
        try: student_answers = json.loads(student_answers)
        except json.JSONDecodeError: student_answers = {}

    # Prepare data for the template (your logic here is good)
    results_data = []
    for question in questions:
        student_answer = student_answers.get(str(question.id), "Not Answered")
        results_data.append(build_question_result(question, student_answer))

    context = {
        'submission': submission,
        'assessment': assessment,
        'results_data': results_data,
    }
    return render(request, 'assessment/view_assessment_result.html', context)


@login_required
def view_exam_result(request, submission_id):
    # --- THE FIX: Look up the ExamSubmission by its ID ---
    submission = get_object_or_404(ExamSubmission.objects.select_related(
        'exam__subject', 'student__user'
    ), id=submission_id)

    # Authorization check
    is_student = request.user == submission.student.user
    is_guardian = hasattr(request.user, 'guardian') and submission.student in request.user.guardian.students.all()

    if not (is_student or is_guardian):
        messages.error(request, "You are not authorized to view these results.")
        return redirect('home')

    exam = submission.exam
    questions = exam.questions.all()
    
    # Defensively handle the 'answers' field
    student_answers = submission.answers
    if isinstance(student_answers, str):
        try: student_answers = json.loads(student_answers)
        except json.JSONDecodeError: student_answers = {}

    # Prepare data for the template
    results_data = []
    for question in questions:
        student_answer = student_answers.get(str(question.id), "Not Answered")
        results_data.append(build_question_result(question, student_answer))

    context = {
        'submission': submission,
        'exam': exam,
        'results_data': results_data,
    }
    return render(request, 'exam/view_exam_result.html', context)